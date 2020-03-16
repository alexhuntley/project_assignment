#!/usr/bin/python3
import openpyxl
import numpy as np
from scipy.optimize import linear_sum_assignment
import tkinter.filedialog, tkinter.scrolledtext
import tkinter as tk
import sys

def build_matrix(ws, keywords, keyword_indices, use_quota=False):
    n_k = len(keyword_indices)
    if use_quota:
        n = sum(row[1].value for row in ws.iter_rows(min_row=2))
    else:
        n = ws.max_row - 1
    M = np.zeros((n, n_k))
    i = 0
    names = []
    for j, row in enumerate(ws.iter_rows(min_row=2)):
        try:
            quota = row[1].value if use_quota else 1
            vals = np.zeros(n_k, dtype=float)
            for k, w in zip(row[2::2], row[3::2]):
                k, w = k.value, w.value
                if k is None or w is None:
                    continue
                vals[keyword_indices[k]] = w            
            M[i:i+quota] = vals
            i += quota
            names.extend([row[0].value]*quota)
        except:
            textbox_print("ERROR in row {} of sheet {}".format(2 + j, ws.title))
            raise
    for i, b in enumerate(np.all(M==0, axis=0)):
        if b:
            textbox_print('Warning: keyword "{}" does not appear in sheet "{}"'.format(
                keywords[i], ws.title))
    return M, names

def textbox_print(string):
    textbox.insert(tk.END, string)
    textbox.insert(tk.END, "\n")

root = tk.Tk()
textbox = tk.scrolledtext.ScrolledText(root)
textbox.pack()

try:
    input_filename = tk.filedialog.askopenfilename(filetypes=[("Excel spreadsheet", "*.xlsx")], parent=root,
                                                   title="Choose input file")
    wb = openpyxl.load_workbook(filename=input_filename)
    keywords_sheet = wb["Keywords"]
    keywords = [row[0].value for row in keywords_sheet.rows]
    keyword_indices = {k: i for i, k in enumerate(keywords)}

    W, repeated_worker_names = build_matrix(wb["Workers"], keywords, keyword_indices, use_quota=True)
    P, project_names = build_matrix(wb["Projects"], keywords, keyword_indices)
    if W.shape[0] < P.shape[0]:
        textbox_print("Warning: number of projects greater than sum of worker quotas; not all projects will be assigned")
    W /= np.linalg.norm(W, axis=1, keepdims=True)
    P /= np.linalg.norm(P, axis=1, keepdims=True)
    C = 1 - W @ P.transpose()
    worker_ind, project_ind = linear_sum_assignment(C)

    output_filename = tk.filedialog.asksaveasfilename(filetypes=[("Excel spreadsheet", "*.xlsx")], parent=root,
                                                      title="Choose output file")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Output"
    out_ws.append(["Worker", "Project", "Score"])
    for w, p in zip(worker_ind, project_ind):
        out_ws.append([repeated_worker_names[w], project_names[p], 1 - C[w,p]])
    out_wb.save(filename=output_filename)
    textbox_print("\nSuccess!\n")
except Exception as e:
    textbox_print("Exception: " + str(e))
    raise
finally:
    root.mainloop()
